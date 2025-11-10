import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

COL_SR_NO = 1
COL_USER_ID = 2
COL_NAME = 6
COL_DEPARTMENT = 17
COL_DESIGNATION = 27

def minutes_to_hhmm(total_minutes):
    if total_minutes is None:
        return "00:00"
    m = int(max(0, round(total_minutes)))
    h, mm = divmod(m, 60)
    return f"{h:02d}:{mm:02d}"

def time_to_minutes(time_str):
    if pd.isna(time_str):
        return None
    s = str(time_str).strip()
    if s in {'', '-'}:
        return None
    try:
        parts = s.split(':')
        if len(parts) == 2:
            hours = int(parts[0]) if parts[0].lstrip('-').isdigit() else 0
            minutes = int(parts[1]) if parts[1].lstrip('-').isdigit() else 0
            sign = -1 if (parts[0].startswith('-') or parts[1].startswith('-')) else 1
            return sign * (abs(hours) * 60 + abs(minutes))
        return None
    except (ValueError, AttributeError):
        return None

def calculate_gross_minutes_from_in_out(first_in, last_out):
    fi = time_to_minutes(first_in)
    lo = time_to_minutes(last_out)
    if fi is None or lo is None:
        return None
    if lo >= fi:
        return lo - fi
    return (24 * 60 - fi) + lo

def classify_attendance(gross_minutes, present_thresh, halfday_thresh, absent_thresh):
    if gross_minutes is None or gross_minutes/60.0 <= absent_thresh:
        return 'Absent'
    gh = gross_minutes / 60.0
    if gh > present_thresh:
        return 'Present'
    if gh > halfday_thresh:
        return 'Half Day'
    return 'Absent'

def safe_to_datetime(val):
    try:
        return pd.to_datetime(val, errors='coerce')
    except Exception:
        return pd.NaT

def get_month_year_from_dates(all_employee_data):
    if all_employee_data:
        dd = all_employee_data[0].get('Daily Data', [])
        for rec in dd:
            dt = safe_to_datetime(rec.get('Date'))
            if not pd.isna(dt):
                return dt.strftime('%B'), dt.year, dt.month
    all_dates = []
    for emp in all_employee_data or []:
        for rec in emp.get('Daily Data', []):
            dt = safe_to_datetime(rec.get('Date'))
            if not pd.isna(dt):
                all_dates.append(dt)
    if all_dates:
        dt = min(all_dates)
        return dt.strftime('%B'), dt.year, dt.month
    return None, None, None

def extract_employee_data(input_file, present_thresh, half_day_thresh, absent_thresh):
    df_raw = pd.read_excel(input_file, sheet_name=0, header=None)
    all_employee_data = []
    n_rows, n_cols = df_raw.shape
    for idx in range(n_rows):
        cell_val = df_raw.iloc[idx, COL_SR_NO] if COL_SR_NO < n_cols else None
        if isinstance(cell_val, (int, float)) and 0 < cell_val < 100:
            sr_no = int(cell_val)
            user_id = df_raw.iloc[idx, COL_USER_ID] if COL_USER_ID < n_cols else ''
            name = df_raw.iloc[idx, COL_NAME] if COL_NAME < n_cols else ''
            department = df_raw.iloc[idx, COL_DEPARTMENT] if COL_DEPARTMENT < n_cols and pd.notna(df_raw.iloc[idx, COL_DEPARTMENT]) else 'N/A'
            designation = df_raw.iloc[idx, COL_DESIGNATION] if COL_DESIGNATION < n_cols and pd.notna(df_raw.iloc[idx, COL_DESIGNATION]) else 'N/A'
            dates_row = None
            for search_idx in range(idx + 1, min(idx + 10, n_rows)):
                val = df_raw.iloc[search_idx, 3] if 3 < n_cols else None
                if isinstance(val, (pd.Timestamp, datetime)) or (pd.notna(safe_to_datetime(val))):
                    dates_row = search_idx
                    break
            if dates_row is None:
                continue
            date_columns = []
            for col_idx in range(3, n_cols):
                dt = safe_to_datetime(df_raw.iloc[dates_row, col_idx])
                if not pd.isna(dt):
                    date_columns.append(col_idx)
            if not date_columns:
                continue
            first_in_row = None
            last_out_row = None
            gross_row = None
            for search_idx in range(dates_row + 1, min(dates_row + 15, n_rows)):
                label = df_raw.iloc[search_idx, 1] if 1 < n_cols else None
                if label == 'First IN':
                    first_in_row = search_idx
                elif label == 'Last OUT':
                    last_out_row = search_idx
                elif label == 'Gross':
                    gross_row = search_idx
            if first_in_row is None or last_out_row is None:
                continue
            daily_data = []
            for col_idx in date_columns:
                date_val = safe_to_datetime(df_raw.iloc[dates_row, col_idx])
                if pd.isna(date_val):
                    continue
                first_in = df_raw.iloc[first_in_row, col_idx] if first_in_row is not None else None
                last_out = df_raw.iloc[last_out_row, col_idx] if last_out_row is not None else None
                gross_time_cell = df_raw.iloc[gross_row, col_idx] if gross_row is not None else None
                gross_minutes = calculate_gross_minutes_from_in_out(first_in, last_out)
                if gross_minutes is None:
                    gm = time_to_minutes(gross_time_cell)
                    gross_minutes = gm if gm is not None else 0
                status = classify_attendance(gross_minutes, present_thresh, half_day_thresh, absent_thresh)
                daily_data.append({
                    'Date': date_val,
                    'First IN': first_in if pd.notna(first_in) else '-',
                    'Last OUT': last_out if pd.notna(last_out) else '-',
                    'Gross Minutes': gross_minutes,
                    'Gross HHMM': minutes_to_hhmm(gross_minutes),
                    'Status': status,
                    'Day of Week': date_val.strftime('%A')
                })
            all_employee_data.append({
                'Sr. No': sr_no,
                'User ID': str(user_id).strip(),
                'Name': str(name).strip(),
                'Department': str(department).strip(),
                'Designation': str(designation).strip(),
                'Daily Data': daily_data
            })
    return all_employee_data

def calculate_summary_statistics(all_employee_data):
    summary_data = []
    for emp in all_employee_data:
        daily_data = emp.get('Daily Data', [])
        working_days_data = [d for d in daily_data if d.get('Day of Week') != 'Sunday']
        present_count = sum(1 for d in working_days_data if d.get('Status') == 'Present')
        half_day_count = sum(1 for d in working_days_data if d.get('Status') == 'Half Day')
        absent_count = sum(1 for d in working_days_data if d.get('Status') == 'Absent')
        total_minutes = sum(int(d.get('Gross Minutes') or 0) for d in working_days_data)
        working_days = present_count + half_day_count
        avg_minutes_per_working_day = int(round(total_minutes / working_days)) if working_days > 0 else 0
        summary_data.append({
            'Sr. No': emp.get('Sr. No'),
            'User ID': emp.get('User ID'),
            'Name': emp.get('Name'),
            'Present Days': present_count,
            'Half Days': half_day_count,
            'Absent Days': absent_count,
            'Total Minutes': total_minutes,
            'Total Hours': minutes_to_hhmm(total_minutes),
            'Avg Minutes/Day': avg_minutes_per_working_day,
            'Avg Hours/Day': minutes_to_hhmm(avg_minutes_per_working_day),
        })
    summary_df = pd.DataFrame(summary_data)
    if not summary_df.empty:
        summary_df = summary_df.drop_duplicates(subset=['User ID'], keep='first').reset_index(drop=True)
        summary_df['Sr. No'] = range(1, len(summary_df) + 1)
    return summary_df

class ExcelStyles:
    @staticmethod
    def get_styles():
        return {
            'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'header_font': Font(bold=True, color="FFFFFF", size=11),
            'title_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'title_font': Font(bold=True, color="FFFFFF", size=14),
            'subheader_fill': PatternFill(start_color="8FC3E6", end_color="8FC3E6", fill_type="solid"),
            'subheader_font': Font(bold=True, size=11),
            'present_fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'halfday_fill': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'absent_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        }

def create_summary_sheet(wb, summary_df, month_name, year):
    styles = ExcelStyles.get_styles()
    ws = wb.create_sheet(f"{month_name} Summary")
    title = f"ATTENDANCE REPORT - {month_name.upper()} {year}"
    ws['A1'] = title
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'] = f"Report Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    headers = ['Sr. No', 'User ID', 'Employee Name',
               'Present Days', 'Half Days', 'Absent Days',
               'Total Hours', 'Avg Hours/Day']
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header_text
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['border']
    for row_idx, (_, emp_row) in enumerate(summary_df.iterrows(), start=6):
        ws.cell(row=row_idx, column=1).value = emp_row['Sr. No']
        ws.cell(row=row_idx, column=2).value = emp_row['User ID']
        ws.cell(row=row_idx, column=3).value = emp_row['Name']
        ws.cell(row=row_idx, column=4).value = int(emp_row['Present Days'])
        ws.cell(row=row_idx, column=5).value = int(emp_row['Half Days'])
        ws.cell(row=row_idx, column=6).value = int(emp_row['Absent Days'])
        ws.cell(row=row_idx, column=7).value = emp_row['Total Hours']
        ws.cell(row=row_idx, column=8).value = emp_row['Avg Hours/Day']
        for c in range(1, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = styles['border']
            if c >= 4:
                cell.alignment = Alignment(horizontal='center')
        if int(emp_row['Present Days']) > 20:
            ws.cell(row=row_idx, column=4).fill = styles['present_fill']
        if int(emp_row['Half Days']) > 0:
            ws.cell(row=row_idx, column=5).fill = styles['halfday_fill']
        if int(emp_row['Absent Days']) > 5:
            ws.cell(row=row_idx, column=6).fill = styles['absent_fill']
    widths = [8, 14, 26, 14, 12, 12, 18, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

def create_daily_attendance_sheet(wb, all_employee_data, month_name, year):
    styles = ExcelStyles.get_styles()
    ws = wb.create_sheet("Daily Attendance")
    title = f"ATTENDANCE REPORT - {month_name.upper()} {year} - DAILY BREAKDOWN"
    ws['A1'] = title
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].font = Font(italic=True, size=10)
    headers_daily = ['Employee ID', 'Employee Name', 'Date', 'First IN', 'Last OUT', 'Gross (HH:MM)', 'Status', 'Day']
    for col_idx, header in enumerate(headers_daily, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['border']
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    row_num = 5
    for emp in all_employee_data:
        for day in emp.get('Daily Data', []):
            dt = safe_to_datetime(day.get('Date'))
            if pd.isna(dt) or dt.weekday() == 6:
                continue
            ws.cell(row=row_num, column=1).value = emp.get('User ID')
            ws.cell(row=row_num, column=2).value = emp.get('Name')
            ws.cell(row=row_num, column=3).value = dt.strftime('%d-%m-%Y')
            ws.cell(row=row_num, column=4).value = day.get('First IN', '-')
            ws.cell(row=row_num, column=5).value = day.get('Last OUT', '-')
            ws.cell(row=row_num, column=6).value = day.get('Gross', '00:00')
            ws.cell(row=row_num, column=7).value = day.get('Status')
            ws.cell(row=row_num, column=8).value = day.get('Day of Week')
            for c in range(1, 9):
                cell = ws.cell(row=row_num, column=c)
                cell.border = styles['border']
                if c in [3, 4, 5, 8]:
                    cell.alignment = Alignment(horizontal='center')
            status = day.get('Status')
            if status == 'Present':
                ws.cell(row=row_num, column=7).fill = styles['present_fill']
            elif status == 'Half Day':
                ws.cell(row=row_num, column=7).fill = styles['halfday_fill']
            else:
                ws.cell(row=row_num, column=7).fill = styles['absent_fill']
            row_num += 1
    ws.auto_filter.ref = f"A4:H{max(row_num-1, 4)}"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    checkin_col = 4
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        cell = row[checkin_col - 1]
        cell_value = cell.value
        if cell_value not in (None, '', '-'):
            try:
                checkin_time = datetime.strptime(str(cell_value), "%H:%M").time()
                if checkin_time > time(9, 45):
                    cell.fill = red_fill
                elif checkin_time > time(9, 30):
                    cell.fill = yellow_fill
            except Exception:
                pass
    widths = [12, 22, 14, 10, 10, 14, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

def create_analysis_sheet(wb, summary_df, month_name, year):
    styles = ExcelStyles.get_styles()
    ws = wb.create_sheet("Analysis & Statistics")
    title = f"ATTENDANCE REPORT - {month_name.upper()} {year} - ANALYSIS"
    ws['A1'] = title
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'] = "OVERALL STATISTICS"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = styles['subheader_fill']
    present_series = pd.to_numeric(summary_df['Present Days'], errors='coerce').fillna(0)
    half_series = pd.to_numeric(summary_df['Half Days'], errors='coerce').fillna(0)
    absent_series = pd.to_numeric(summary_df['Absent Days'], errors='coerce').fillna(0)
    total_minutes_series = pd.to_numeric(summary_df['Total Minutes'], errors='coerce').fillna(0)
    total_present = int(present_series.sum())
    total_half = int(half_series.sum())
    total_absent = int(absent_series.sum())
    total_working_days = total_present + total_half + total_absent
    stats = [
        ['Total Employees', len(summary_df)],
        ['Total Working Days (Non-Sundays)', total_working_days],
        ['Average Days per Employee', round(total_working_days / len(summary_df), 2) if len(summary_df) > 0 else 0.0],
    ]
    row = 4
    for label, value in stats:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).number_format = '0.00'
        row += 1
    ws['A10'] = "ATTENDANCE BREAKDOWN"
    ws['A10'].font = Font(bold=True, size=12)
    ws['A10'].fill = styles['subheader_fill']
    headers_break = ['Status', 'Total Days', 'Percentage', 'Avg Per Employee']
    for col_idx, header in enumerate(headers_break, start=1):
        cell = ws.cell(row=11, column=col_idx)
        cell.value = header
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['border']
    pct = lambda x: round((x/total_working_days)*100, 2) if total_working_days > 0 else 0.0
    breakdown = [
        ['Present', total_present, pct(total_present), round(total_present/len(summary_df), 2) if len(summary_df) > 0 else 0.0],
        ['Half Day', total_half, pct(total_half), round(total_half/len(summary_df), 2) if len(summary_df) > 0 else 0.0],
        ['Absent', total_absent, pct(total_absent), round(total_absent/len(summary_df), 2) if len(summary_df) > 0 else 0.0],
    ]
    row = 12
    for status, days, pc, avg in breakdown:
        ws.cell(row=row, column=1).value = status
        ws.cell(row=row, column=2).value = days
        ws.cell(row=row, column=3).value = f"{pc}%"
        ws.cell(row=row, column=4).value = avg
        fill = styles['present_fill'] if 'Present' in status else styles['halfday_fill'] if 'Half' in status else styles['absent_fill']
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.border = styles['border']
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            if c >= 4:
                cell.number_format = '0.00'
        row += 1
    ws['A17'] = "HOURS ANALYSIS"
    ws['A17'].font = Font(bold=True, size=12)
    ws['A17'].fill = styles['subheader_fill']
    hours_data = [
        ['Total Hours Worked', minutes_to_hhmm(int(total_minutes_series.sum()))],
        ['Average Hours per Employee', minutes_to_hhmm(int(round(total_minutes_series.sum() / len(summary_df))) if len(summary_df) > 0 else 0)],
        ['Maximum Hours (Single Employee)', minutes_to_hhmm(int(total_minutes_series.max()) if len(total_minutes_series) > 0 else 0)],
        ['Minimum Hours (Single Employee)', minutes_to_hhmm(int(total_minutes_series.min()) if len(total_minutes_series) > 0 else 0)],
        ['Average Hours per Working Day', minutes_to_hhmm(int(pd.to_numeric(summary_df['Avg Minutes/Day'], errors='coerce').fillna(0).mean())) if 'Avg Minutes/Day' in summary_df.columns else '00:00'],
    ]
    row = 18
    for label, value in hours_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 1
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 22

############### Streamlit UI #####################

def time_to_hours(t: time):
    return t.hour + t.minute / 60.0

st.markdown(
    """
    <h1 style='text-align: center; color: #4B8BBE; font-weight: bold;
    font-family: "Segoe UI", Tahoma, Geneva, Verdana, sans-serif;'>BOW AND BAAN ATTENDANCE</h1>
    <hr style='border: 1px solid #BBB;'/>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("Attendance Settings")
    present_time = st.time_input("Present Threshold (HH:MM)", value=time(6, 0))
    half_day_time = st.time_input("Half Day Threshold (HH:MM)", value=time(0, 0))
    absent_time = st.time_input("Absent Threshold (HH:MM)", value=time(0, 0))

present_threshold = time_to_hours(present_time)
half_day_threshold = time_to_hours(half_day_time)
absent_threshold = time_to_hours(absent_time)

uploaded_file = st.file_uploader("Upload your attendance Excel file", type=["xlsx"])

if uploaded_file:
    with st.spinner("Generating report..."):
        input_excel = io.BytesIO(uploaded_file.read())
        output = io.BytesIO()
        all_employee_data = extract_employee_data(input_excel, present_threshold, half_day_threshold, absent_threshold)
        if not all_employee_data:
            st.error("No employee data found in input file.")
        else:
            month_name, year, _ = get_month_year_from_dates(all_employee_data)
            summary_df = calculate_summary_statistics(all_employee_data)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            create_summary_sheet(wb, summary_df, month_name, year)
            create_daily_attendance_sheet(wb, all_employee_data, month_name, year)
            create_analysis_sheet(wb, summary_df, month_name, year)
            wb.save(output)
            output.seek(0)
            st.success("Report generated successfully!")
            st.download_button(
                label="Download Attendance Report",
                data=output,
                file_name=f"Attendance_Report_{month_name}_{year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
