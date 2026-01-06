import streamlit as st
import pandas as pd
from datetime import datetime, time, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
import io

# Constants for Excel columns (update these indexes based on your Excel structure)
COL_SR_NO = 1
COL_USER_ID = 2
COL_NAME = 6
COL_DEPARTMENT = 17
COL_DESIGNATION = 27

def minutes_to_hhmm(total_minutes):
    if total_minutes is None:
        return "00:00"
    m = int(max(0, round(total_minutes)))
    h, mm = divmod(m, 60)
    return f"{h:02d}:{mm:02d}"

def time_to_minutes(time_str):
    if pd.isna(time_str):
        return None
    s = str(time_str).strip()
    if s in {'', '-'}:
        return None
    try:
        parts = s.split(':')
        if len(parts) == 2:
            hours = int(parts[0]) if parts[0].lstrip('-').isdigit() else 0
            minutes = int(parts[1]) if parts[1].lstrip('-').isdigit() else 0
            sign = -1 if (parts[0].startswith('-') or parts[1].startswith('-')) else 1
            return sign * (abs(hours) * 60 + abs(minutes))
        return None
    except (ValueError, AttributeError):
        return None

def calculate_gross_minutes_from_in_out(first_in, last_out):
    fi = time_to_minutes(first_in)
    lo = time_to_minutes(last_out)
    if fi is None or lo is None:
        return None
    if lo >= fi:
        return lo - fi
    return (24 * 60 - fi) + lo

def classify_attendance(gross_minutes, present_thresh, halfday_thresh, employee_name,
                        sat_working_employees, gov_holidays, date_val):
    date_only = date_val.date()
    if date_only in gov_holidays:
        return 'Holiday'
    # Saturday Logic: Default is Holiday. If employee is in 'Working' list, they work.
    if date_val.weekday() == 5:
        if employee_name not in sat_working_employees:
            return 'Holiday'
    if gross_minutes is None or gross_minutes <= 0:
        return 'Absent'
    gh = gross_minutes / 60.0
    if gh > present_thresh:
        return 'Present'
    if gh > halfday_thresh:
        return 'Half Day'
    return 'Absent'

def safe_to_datetime(val):
    try:
        return pd.to_datetime(val, errors='coerce')
    except Exception:
        return pd.NaT

def time_to_hours(t: time):
    return t.hour + t.minute / 60.0

def get_month_year_from_dates(all_employee_data):
    if all_employee_data:
        dd = all_employee_data[0].get('Daily Data', [])
        for rec in dd:
            dt = safe_to_datetime(rec.get('Date'))
            if not pd.isna(dt):
                return dt.strftime('%B'), dt.year, dt.month
    all_dates = []
    for emp in all_employee_data or []:
        for rec in emp.get('Daily Data', []):
            dt = safe_to_datetime(rec.get('Date'))
            if not pd.isna(dt):
                all_dates.append(dt)
    if all_dates:
        dt = min(all_dates)
        return dt.strftime('%B'), dt.year, dt.month
    return None, None, None

def extract_employee_data(input_file, present_thresh, half_day_thresh,
                          sat_working_employees, gov_holidays):
    df_raw = pd.read_excel(input_file, sheet_name=0, header=None, engine="openpyxl")
    all_employee_data = []
    n_rows, n_cols = df_raw.shape
    for idx in range(n_rows):
        cell_val = df_raw.iloc[idx, COL_SR_NO] if COL_SR_NO < n_cols else None
        if isinstance(cell_val, (int, float)) and 0 < cell_val < 100:
            sr_no = int(cell_val)
            user_id = df_raw.iloc[idx, COL_USER_ID] if COL_USER_ID < n_cols else ''
            name = df_raw.iloc[idx, COL_NAME] if COL_NAME < n_cols else ''
            department = df_raw.iloc[idx, COL_DEPARTMENT] if COL_DEPARTMENT < n_cols and pd.notna(df_raw.iloc[idx, COL_DEPARTMENT]) else 'N/A'
            designation = df_raw.iloc[idx, COL_DESIGNATION] if COL_DESIGNATION < n_cols and pd.notna(df_raw.iloc[idx, COL_DESIGNATION]) else 'N/A'
            dates_row = None
            for search_idx in range(idx + 1, min(idx + 10, n_rows)):
                val = df_raw.iloc[search_idx, 3] if 3 < n_cols else None
                if isinstance(val, (pd.Timestamp, datetime)) or (pd.notna(safe_to_datetime(val))):
                    dates_row = search_idx
                    break
            if dates_row is None:
                continue
            date_columns = []
            for col_idx in range(3, n_cols):
                dt = safe_to_datetime(df_raw.iloc[dates_row, col_idx])
                if not pd.isna(dt):
                    date_columns.append(col_idx)
            if not date_columns:
                continue
            first_in_row = None
            last_out_row = None
            gross_row = None
            for search_idx in range(dates_row + 1, min(dates_row + 15, n_rows)):
                label = df_raw.iloc[search_idx, 1] if 1 < n_cols else None
                if label == 'First IN':
                    first_in_row = search_idx
                elif label == 'Last OUT':
                    last_out_row = search_idx
                elif label == 'Gross':
                    gross_row = search_idx
            if first_in_row is None or last_out_row is None:
                continue
            daily_data = []
            for col_idx in date_columns:
                date_val = safe_to_datetime(df_raw.iloc[dates_row, col_idx])
                if pd.isna(date_val):
                    continue
                first_in = df_raw.iloc[first_in_row, col_idx] if first_in_row is not None else None
                last_out = df_raw.iloc[last_out_row, col_idx] if last_out_row is not None else None
                gross_time_cell = df_raw.iloc[gross_row, col_idx] if gross_row is not None else None
                gross_minutes = calculate_gross_minutes_from_in_out(first_in, last_out)
                if gross_minutes is None:
                    gm = time_to_minutes(gross_time_cell)
                    gross_minutes = gm if gm is not None else 0
                status = classify_attendance(gross_minutes, present_thresh, half_day_thresh,
                                             name, sat_working_employees,
                                             gov_holidays, date_val)
                daily_data.append({
                    'Date': date_val,
                    'First IN': first_in if pd.notna(first_in) else '-',
                    'Last OUT': last_out if pd.notna(last_out) else '-',
                    'Gross Minutes': gross_minutes,
                    'Gross HHMM': minutes_to_hhmm(gross_minutes),
                    'Status': status,
                    'Day of Week': date_val.strftime('%A')
                })
            all_employee_data.append({
                'Sr. No': sr_no,
                'User ID': str(user_id).strip(),
                'Name': str(name).strip(),
                'Department': str(department).strip(),
                'Designation': str(designation).strip(),
                'Daily Data': daily_data
            })
    return all_employee_data


def calculate_summary_statistics(all_employee_data):
    summary_data = []
    for emp in all_employee_data:
        daily_data = emp.get('Daily Data', [])
        working_days = [d for d in daily_data if d.get('Day of Week') != 'Sunday' and d.get('Status') != 'Holiday']
        present_count = sum(1 for d in working_days if d.get('Status') == 'Present')
        half_day_count = sum(1 for d in working_days if d.get('Status') == 'Half Day')
        absent_count = sum(1 for d in working_days if d.get('Status') == 'Absent')
        total_minutes = sum(int(d.get('Gross Minutes') or 0) for d in working_days)
        working_day_count = present_count + half_day_count
        avg_minutes_per_day = int(round(total_minutes / working_day_count)) if working_day_count > 0 else 0
        summary_data.append({
            'Sr. No': emp.get('Sr. No'),
            'User ID': emp.get('User ID'),
            'Name': emp.get('Name'),
            'Present Days': present_count,
            'Half Days': half_day_count,
            'Absent Days': absent_count,
            'Total Minutes': total_minutes,
            'Total Hours': minutes_to_hhmm(total_minutes),
            'Avg Minutes/Day': avg_minutes_per_day,
            'Avg Hours/Day': minutes_to_hhmm(avg_minutes_per_day),
        })
    summary_df = pd.DataFrame(summary_data)
    if not summary_df.empty:
        summary_df = summary_df.drop_duplicates(subset=['User ID'], keep='first').reset_index(drop=True)
        summary_df['Sr. No'] = range(1, len(summary_df) + 1)
    return summary_df


class ExcelStyles:
    @staticmethod
    def get_styles():
        return {
            'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'header_font': Font(bold=True, color="FFFFFF", size=11),
            'title_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'title_font': Font(bold=True, color="FFFFFF", size=14),
            'subheader_fill': PatternFill(start_color="8FC3E6", end_color="8FC3E6", fill_type="solid"),
            'subheader_font': Font(bold=True, size=11),
            'present_fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'halfday_fill': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'absent_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'holiday_fill': PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid"),
            'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        }


def create_summary_sheet(wb, summary_df, month_name, year):
    styles = ExcelStyles.get_styles()
    ws = wb.create_sheet(f"{month_name} Summary")
    title = f"ATTENDANCE REPORT - {month_name.upper()} {year}"
    ws['A1'] = title
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'] = f"Report Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    headers = ['Sr. No', 'User ID', 'Employee Name',
               'Present Days', 'Half Days', 'Absent Days',
               'Total Hours (HH:MM)', 'Avg Hours/Day (HH:MM)']
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header_text
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['border']
    for row_idx, (_, emp_row) in enumerate(summary_df.iterrows(), start=6):
        ws.cell(row=row_idx, column=1).value = emp_row['Sr. No']
        ws.cell(row=row_idx, column=2).value = emp_row['User ID']
        ws.cell(row=row_idx, column=3).value = emp_row['Name']
        ws.cell(row=row_idx, column=4).value = int(emp_row['Present Days'])
        ws.cell(row=row_idx, column=5).value = int(emp_row['Half Days'])
        ws.cell(row=row_idx, column=6).value = int(emp_row['Absent Days'])
        ws.cell(row=row_idx, column=7).value = emp_row['Total Hours']
        ws.cell(row=row_idx, column=8).value = emp_row['Avg Hours/Day']
        for c in range(1, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = styles['border']
            if c >= 4:
                cell.alignment = Alignment(horizontal='center')
        if int(emp_row['Present Days']) > 20:
            ws.cell(row=row_idx, column=4).fill = styles['present_fill']
        if int(emp_row['Half Days']) > 0:
            ws.cell(row=row_idx, column=5).fill = styles['halfday_fill']
        if int(emp_row['Absent Days']) > 5:
            ws.cell(row=row_idx, column=6).fill = styles['absent_fill']
    widths = [8, 14, 26, 14, 12, 12, 18, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width


def create_daily_attendance_sheet(wb, all_employee_data, month_name, year):
    styles = ExcelStyles.get_styles()
    ws = wb.create_sheet("Daily Attendance")
    title = f"ATTENDANCE REPORT - {month_name.upper()} {year} - DAILY BREAKDOWN"
    ws['A1'] = title
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    headers_daily = ['Employee ID', 'Employee Name', 'Date', 'First IN', 'Last OUT',
                     'Gross (HH:MM)', 'Status', 'Day']
    for col_idx, header in enumerate(headers_daily, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['border']
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    row_num = 5
    for emp in all_employee_data:
        for day in emp.get('Daily Data', []):
            dt = safe_to_datetime(day.get('Date'))
            if pd.isna(dt) or dt.weekday() == 6:
                continue
            ws.cell(row=row_num, column=1).value = emp.get('User ID')
            ws.cell(row=row_num, column=2).value = emp.get('Name')
            ws.cell(row=row_num, column=3).value = dt.strftime('%d-%m-%Y')
            ws.cell(row=row_num, column=4).value = day.get('First IN', '-')
            ws.cell(row=row_num, column=5).value = day.get('Last OUT', '-')
            ws.cell(row=row_num, column=6).value = day.get('Gross HHMM', '00:00')
            ws.cell(row=row_num, column=7).value = day.get('Status')
            ws.cell(row=row_num, column=8).value = day.get('Day of Week')
            for c in range(1, 9):
                cell = ws.cell(row=row_num, column=c)
                cell.border = styles['border']
                if c in [3, 4, 5, 8]:
                    cell.alignment = Alignment(horizontal='center')
            status = day.get('Status')
            if status == 'Present':
                ws.cell(row=row_num, column=7).fill = styles['present_fill']
            elif status == 'Half Day':
                ws.cell(row=row_num, column=7).fill = styles['halfday_fill']
            elif status == 'Holiday':
                ws.cell(row=row_num, column=7).fill = styles['holiday_fill']
            else:
                ws.cell(row=row_num, column=7).fill = styles['absent_fill']
            row_num += 1
    ws.auto_filter.ref = f"A4:H{max(row_num-1, 4)}"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    checkin_col = 4
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        cell = row[checkin_col - 1]
        cell_value = cell.value
        if cell_value not in (None, '', '-'):
            try:
                checkin_time = datetime.strptime(str(cell_value), "%H:%M").time()
                if checkin_time > time(9, 45):
                    cell.fill = red_fill
                elif checkin_time > time(9, 30):
                    cell.fill = yellow_fill
            except Exception:
                pass
    widths = [12, 22, 14, 10, 10, 14, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width



def create_analysis_sheet(wb, summary_df, month_name, year):
    styles = ExcelStyles.get_styles()
    ws = wb.create_sheet("Analysis & Statistics")
    
    # --- Title ---
    title = f"ATTENDANCE REPORT - {month_name.upper()} {year} - ANALYSIS"
    ws['A1'] = title
    ws['A1'].font = styles['title_font']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Basic Series
    present_series = pd.to_numeric(summary_df['Present Days'], errors='coerce').fillna(0)
    half_series = pd.to_numeric(summary_df['Half Days'], errors='coerce').fillna(0)
    absent_series = pd.to_numeric(summary_df['Absent Days'], errors='coerce').fillna(0)
    total_minutes_series = pd.to_numeric(summary_df['Total Minutes'], errors='coerce').fillna(0)
    
    # Calculations
    total_present = int(present_series.sum())
    total_half = int(half_series.sum())
    total_absent = int(absent_series.sum())
    total_working_days = total_present + total_half
    
    # --- 1. ATTENDANCE BREAKDOWN ---
    row = 3
    headers = ["ATTENDANCE BREAKDOWN", "", ""]
    ws.cell(row=row, column=1).value = headers[0]
    ws.cell(row=row, column=1).font = styles['subheader_font']
    ws.cell(row=row, column=1).fill = styles['subheader_fill']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    
    row += 1
    # Table Header
    ws.cell(row=row, column=1).value = "Metric"
    ws.cell(row=row, column=2).value = "Count"
    ws.cell(row=row, column=3).value = "Formula"
    for c in range(1, 4):
        ws.cell(row=row, column=c).font = styles['header_font']
        ws.cell(row=row, column=c).fill = styles['header_fill']
        ws.cell(row=row, column=c).border = styles['border']

    row += 1
    breakdown_data = [
        ("Total Present Days", total_present, "SUM of Present Days column"),
        ("Total Half Days", total_half, "SUM of Half Days column"),
        ("Total Absent Days", total_absent, "SUM of Absent Days column"),
        ("Total Working Days", total_working_days, "Present Days + Half Days"),
    ]

    for metric, count, formula in breakdown_data:
        ws.cell(row=row, column=1).value = metric
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = formula
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = styles['border']
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 1
        
    # --- 2. HOURS ANALYSIS ---
    row += 1
    ws.cell(row=row, column=1).value = "HOURS ANALYSIS"
    ws.cell(row=row, column=1).font = styles['subheader_font']
    ws.cell(row=row, column=1).fill = styles['subheader_fill']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    
    row += 1
    ws.cell(row=row, column=1).value = "Metric"
    ws.cell(row=row, column=2).value = "Value (HH:MM)"
    ws.cell(row=row, column=3).value = "Formula"
    for c in range(1, 4):
        ws.cell(row=row, column=c).font = styles['header_font']
        ws.cell(row=row, column=c).fill = styles['header_fill']
        ws.cell(row=row, column=c).border = styles['border']
        
    total_hours_worked_mins = int(total_minutes_series.sum())
    
    # Avg hours per employee per day
    # Formula: Average of the 'Avg Minutes/Day' column from summary
    avg_mins_per_day_series = pd.to_numeric(summary_df['Avg Minutes/Day'], errors='coerce').fillna(0)
    avg_mins_emp_day = int(avg_mins_per_day_series.mean()) if not avg_mins_per_day_series.empty else 0
    
    # Max/Min Employee
    max_mins = total_minutes_series.max()
    min_mins = total_minutes_series.min()
    
    max_emp_name = summary_df.loc[total_minutes_series.idxmax(), 'Name'] if not summary_df.empty else "N/A"
    min_emp_name = summary_df.loc[total_minutes_series.idxmin(), 'Name'] if not summary_df.empty else "N/A"

    row += 1
    hours_data = [
        ("Total Hours Worked (All Employees)", minutes_to_hhmm(total_hours_worked_mins), "SUM of Total Minutes / 60"),
        ("Average Hours per Employee per Day", minutes_to_hhmm(avg_mins_emp_day), "AVERAGE(Avg Hours/Day column)"),
        (f"Max Hours (Employee: {max_emp_name})", minutes_to_hhmm(max_mins), ""),
        (f"Min Hours (Employee: {min_emp_name})", minutes_to_hhmm(min_mins), "")
    ]
    
    for metric, val, formula in hours_data:
        ws.cell(row=row, column=1).value = metric
        ws.cell(row=row, column=2).value = val
        ws.cell(row=row, column=3).value = formula
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = styles['border']
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 1

    # --- 3. OVERALL METRICS ---
    row += 1
    ws.cell(row=row, column=1).value = "OVERALL METRICS"
    ws.cell(row=row, column=1).font = styles['subheader_font']
    ws.cell(row=row, column=1).fill = styles['subheader_fill']
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    row += 1
    ws.cell(row=row, column=1).value = "Metric"
    ws.cell(row=row, column=2).value = "Value"
    ws.cell(row=row, column=3).value = "Formula"
    for c in range(1, 4):
        ws.cell(row=row, column=c).font = styles['header_font']
        ws.cell(row=row, column=c).fill = styles['header_fill']
        ws.cell(row=row, column=c).border = styles['border']
        
    total_employees = len(summary_df)
    total_days_all = total_working_days + total_absent
    
    if total_days_all > 0:
        overall_attendance_rate = (total_working_days / total_days_all) * 100
        half_day_pct = (total_half / total_days_all) * 100
    else:
        overall_attendance_rate = 0.0
        half_day_pct = 0.0

    row += 1
    metrics_data = [
        ("Total Employees", total_employees, "COUNT of unique User IDs"),
        ("Overall Attendance Rate (%)", f"{overall_attendance_rate:.2f}%", "(Working Days / (Working Days + Absent Days)) * 100"),
        ("Half Day Percentage (%)", f"{half_day_pct:.2f}%", "(Half Days / Total Days) * 100")
    ]
    
    for metric, val, formula in metrics_data:
        ws.cell(row=row, column=1).value = metric
        ws.cell(row=row, column=2).value = val
        ws.cell(row=row, column=3).value = formula
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = styles['border']
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        row += 1
        
    # Column Widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45

    # --- 4. BAR CHART (Employee Work Hours) ---
    # Chart removed as per user request
    pass


# Streamlit UI Logic
st.set_page_config(page_title="Attendance Report", layout="wide", initial_sidebar_state="expanded")
st.markdown(
    """
    <h1 style='text-align: center; color: #4B8BBE; font-weight: bold;
    font-family: "Segoe UI", Tahoma, Geneva, Verdana, sans-serif;'>BOW AND BAAN ATTENDANCE</h1>
    <hr style='border: 1px solid #BBB;'/>
    """, unsafe_allow_html=True)

uploaded_file = st.file_uploader("Upload your attendance Excel file", type=["xlsx"])

with st.sidebar:
    st.header("Settings")
    present_time = st.time_input("Present Threshold (Clock)", value=time(8, 0))
    half_day_time = st.time_input("Half Day Threshold (Clock)", value=time(4, 0))

    departments = []
    designations = []
    employee_names = []
    if uploaded_file:
        # Use uploaded_file directly in read_excel without re-wrapping
        df_raw = pd.read_excel(uploaded_file, header=None, engine="openpyxl")
        departments = df_raw[COL_DEPARTMENT].dropna().unique().tolist()
        designations = df_raw[COL_DESIGNATION].dropna().unique().tolist()
        employee_names = df_raw[COL_NAME].dropna().unique().tolist()
        st.session_state['departments'] = departments
        st.session_state['designations'] = designations
        st.session_state['employee_names'] = employee_names
    else:
        departments = st.session_state.get('departments', [])
        designations = st.session_state.get('designations', [])
        employee_names = st.session_state.get('employee_names', [])

    sat_working_employees = st.multiselect("Employees working on saturdays:", 
                                           employee_names, key="sat_working_emp")

    st.markdown("#### :calendar: Government Holidays")
    if "gov_holiday_dates" not in st.session_state:
        st.session_state["gov_holiday_dates"] = []
    new_holiday = st.date_input("Select a holiday date", value=date.today())
    add_holiday = st.button("➕ Add Holiday")
    clear_holidays = st.button("🗑️ Clear All")
    if add_holiday and new_holiday and new_holiday not in st.session_state["gov_holiday_dates"]:
        st.session_state["gov_holiday_dates"].append(new_holiday)
    if clear_holidays:
        st.session_state["gov_holiday_dates"] = []

    gov_holidays = set(st.session_state["gov_holiday_dates"])

    st.write("Selected Holidays:", sorted(str(d) for d in gov_holidays))


present_threshold = time_to_hours(present_time)
half_day_threshold = time_to_hours(half_day_time)


if uploaded_file and st.button("Generate Attendance Report"):
    with st.spinner("Generating report..."):
        all_employee_data = extract_employee_data(
            uploaded_file,
            present_threshold,
            half_day_threshold,
            st.session_state.get('sat_working_emp', []),
            gov_holidays
        )
        if not all_employee_data:
            st.error("No employee data found in input file.")
        else:
            month_name, year, _ = get_month_year_from_dates(all_employee_data)
            summary_df = calculate_summary_statistics(all_employee_data)
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            create_summary_sheet(wb, summary_df, month_name, year)
            create_daily_attendance_sheet(wb, all_employee_data, month_name, year)
            create_analysis_sheet(wb, summary_df, month_name, year)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            st.success("Report generated successfully!")
            st.download_button(
                label="Download Attendance Report",
                data=output,
                file_name=f"Attendance_Report_{month_name}_{year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
